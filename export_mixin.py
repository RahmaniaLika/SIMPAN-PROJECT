from datetime import datetime
import os


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportMixin:
    def _get_export_data(self):
        """Override method ini di class yang inherit dari ExportMixin"""
        raise NotImplementedError("_get_export_data() belum diimplementasikan")

    def _buat_nama_file(self, prefix, extension):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
    
    def _pastikan_folder_export(self, folder="exports"):
        if not os.path.exists(folder):
            os.makedirs(folder) 
            print(f" Folder '{folder}' telah dibuat")
        return folder
    
    def export_to_excel(self, nama_file=None, folder="exports"):
        try:
            folder_path = self._pastikan_folder_export(folder)
            
            if nama_file is None:
                nama_file = self._buat_nama_file("laporan_simpan", "xlsx")

            full_path = os.path.join(folder_path, nama_file)

            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Barang"


            data = self._get_export_data()
            
            header_fill = PatternFill(
                start_color="366092",
                end_color="366092", 
                fill_type="solid"
            )

            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )


            headers = data['headers']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                cell.border = border


            for row_idx, row_data in enumerate(data['rows'], 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                    
                    if col_idx in data.get('currency_columns', []):
                        cell.number_format = '#,##0'
            

            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            

            wb.save(full_path)
            
            print(f" Excel berhasil dibuat: {full_path}")
            return full_path
            
        except Exception as e:
            print(f" Error export Excel: {e}")
            raise